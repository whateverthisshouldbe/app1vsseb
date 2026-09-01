#!/usr/bin/env python3
"""Genereert schemaData.js en winkelData.js uit het Excel-bronbestand.

Gebruik:  python3 tools/build-data.py <pad naar Seb60dagenvoedingsschema.xlsx>

De app leest de gegenereerde modules via dynamische imports in index.html.
Prijzen en winkelindeling staan in tools/winkels.json, want die staan niet
in het Excel-bestand.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WINKELS = ROOT / "tools" / "winkels.json"
BEREIDING = ROOT / "tools" / "bereiding.json"
FOTOS = ROOT / "fotos"


def tekst(v):
    return "" if v is None else str(v).strip()


def getal(v):
    if v is None or v == "":
        return 0
    try:
        return round(float(v), 1)
    except (TypeError, ValueError):
        return 0


def lees_recepten(ws, prefix_strip=True):
    """Leest een receptenblad: Nr | Gerecht | Ingredient | Gram | kcal | E | V | K."""
    recepten = {}
    huidig = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        nr, gerecht, ing = tekst(r[0]), tekst(r[1]), tekst(r[2])
        if nr:
            huidig = nr
            recepten[nr] = {"naam": gerecht, "ing": [], "tot": {}}
        if not huidig or not ing:
            continue
        rec = recepten[huidig]
        if ing.lower() == "portie totaal":
            rec["tot"] = {
                "kcal": getal(r[4]), "e": getal(r[5]),
                "v": getal(r[6]), "k": getal(r[7]),
            }
        else:
            rec["ing"].append({
                # 'g' is wat het maaltijdscherm uitleest, 'q' de rest
                "n": ing, "q": getal(r[3]), "g": getal(r[3]), "eh": "g",
                "kcal": getal(r[4]), "e": getal(r[5]),
                "v": getal(r[6]), "k": getal(r[7]),
            })
    return recepten


def lees_dagen(wb):
    """Combineert '60-dagen schema' (dagregels) met 'Dagdetails' (blokken)."""
    blokken = {}
    totalen = {}
    for r in wb["Dagdetails"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        d, blok = int(r[0]), tekst(r[2])
        macro = {
            "kcal": getal(r[4]), "e": getal(r[5]),
            "v": getal(r[6]), "k": getal(r[7]), "vezel": getal(r[8]),
        }
        if blok == "TOTAAL":
            totalen[d] = macro
            continue
        wat = tekst(r[3])
        if not wat or wat == "-":
            continue
        blokken.setdefault(d, []).append({"blok": blok, "wat": wat, **macro})

    dagen = []
    for r in wb["60-dagen schema"].iter_rows(min_row=2, values_only=True):
        if not isinstance(r[0], (int, float)):
            continue  # slaat de 'Gemiddeld'-regel onderaan over
        d = int(r[0])
        tot = totalen.get(d, {})
        dagen.append({
            "d": d,
            "wd": tekst(r[1]),
            "tr": tekst(r[2]),
            "lunch": tekst(r[5]),
            "diner": tekst(r[7]),
            "kcal": tot.get("kcal", getal(r[10])),
            "e": tot.get("e", getal(r[11])),
            "v": tot.get("v", getal(r[12])),
            "k": tot.get("k", getal(r[13])),
            "vezel": tot.get("vezel", getal(r[14])),
            "blokken": blokken.get(d, []),
        })
    return dagen


def lees_supplementen(ws):
    """Het dagprotocol staat onder de kop 'Middel'; daarna volgt een tweede tabel."""
    supp, aandacht, sectie = [], [], None
    for r in ws.iter_rows(values_only=True):
        cellen = [tekst(c) for c in r]
        eerste = next((c for c in cellen if c), "")
        if eerste == "Middel":
            sectie = "protocol"
            continue
        if eerste == "Dag":
            sectie = "aandacht"
            continue
        if "DAGEN DIE EXTRA" in eerste:
            continue
        rij = [c for c in cellen[1:] if c]
        if sectie == "protocol" and len(rij) >= 4:
            supp.append({
                "naam": rij[0], "dosis": rij[1],
                "wanneer": rij[2], "waarom": rij[3],
                "moment": moment_van(rij[0], rij[2]),
            })
        elif sectie == "aandacht" and len(rij) >= 3 and rij[0].isdigit():
            aandacht.append({
                "d": int(rij[0]), "wd": rij[1],
                "tekort": rij[2], "fix": rij[3] if len(rij) > 3 else "",
            })
    return supp, aandacht


def ratio(v):
    """Verhoudingen hebben meer decimalen nodig dan getal() geeft."""
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return 0.0


def lees_micros(ws):
    """Rij met 'Referentie' geeft de norm; daarna per dag de absolute waarden."""
    rijen = list(ws.iter_rows(values_only=True))
    ref_rij = next(r for r in rijen if tekst(r[0]) == "Referentie")
    kop = next(r for r in rijen if tekst(r[0]) == "Dag")
    namen = [tekst(c) for c in kop[2:] if tekst(c)]
    ref = [ratio(c) for c in ref_rij[2:2 + len(namen)]]
    per_dag = {}
    for r in rijen:
        if not isinstance(r[0], (int, float)):
            continue
        per_dag[int(r[0])] = [ratio(c) for c in r[2:2 + len(namen)]]
    return {"namen": namen, "ref": ref, "perDag": per_dag}


# Waar een supplement in de dag valt; 'Losse dagen' hoort niet bij een
# vast moment en telt dus niet mee in het dagelijkse rijtje.
MOMENTEN = [
    ("bij het ontbijt", "Ontbijt"),
    ("bij een maaltijd", "Lunch"),
    ("bij het eten", "Diner"),
    ("timing maakt niet uit", "Ontbijt"),
]

# Deze horen niet in het dagelijkse afvinkrijtje: keukenzout is een
# kookinstructie, D-Bloat is uitdrukkelijk niet dagelijks.
GEEN_MOMENT = ("keukenzout", "d-bloat")


def moment_van(naam, wanneer):
    if any(x in naam.lower() for x in GEEN_MOMENT):
        return ""
    w = wanneer.lower()
    for sleutel, moment in MOMENTEN:
        if sleutel in w:
            return moment
    return ""


def lees_boodschappen(ws):
    weken, week, dagen = [], None, ""
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is not None:
            week, dagen = int(r[0]), tekst(r[1])
            weken.append({"week": week, "dagen": dagen, "items": []})
        naam = tekst(r[2])
        # 'einde week N' zijn scheidingsregels, geen boodschappen
        if not naam or naam.lower().startswith("einde week") or not weken:
            continue
        weken[-1]["items"].append({
            "n": naam, "q": getal(r[3]),
            "eh": tekst(r[4]), "opm": tekst(r[5]),
        })
    return weken


def hernoem_items(weken, hernoem):
    """Past de naamwijzigingen toe en telt dubbelen binnen een week op.

    Na het hernoemen kan hetzelfde product twee keer in een week staan
    (rundergehakt 5% en 15% worden allebei 'rundergehakt'), dus die regels
    moeten samengevoegd worden.
    """
    # gram en kilo (en liter) van hetzelfde product zijn dezelfde regel
    naar_basis = {"kg": ("g", 1000), "l": ("liter", 1)}
    for w in weken:
        samen = {}
        for i in w["items"]:
            i["n"] = hernoem.get(i["n"], i["n"])
            eh, factor = naar_basis.get(i["eh"], (i["eh"], 1))
            i["eh"], i["q"] = eh, i["q"] * factor
            sleutel = (i["n"], eh)
            if sleutel in samen:
                samen[sleutel]["q"] = round(samen[sleutel]["q"] + i["q"], 1)
                # de opmerking van de eerste regel blijft staan
            else:
                samen[sleutel] = i
        for i in samen.values():
            if i["eh"] == "g" and i["q"] >= 1000:
                i["eh"], i["q"] = "kg", round(i["q"] / 1000, 2)
            else:
                i["q"] = round(i["q"], 1)
        w["items"] = list(samen.values())


def hernoem_recepten(recept, hernoem):
    for r in recept.values():
        for i in r["ing"]:
            i["n"] = hernoem.get(i["n"], i["n"])


def naar_basis(q, eh):
    """Rekent een regel om naar kilo of liter; None als dat niet kan."""
    e = eh.lower()
    if e == "kg":
        return q
    if e == "g":
        return q / 1000
    if e in ("liter", "l"):
        return q
    return None


def bulk_verdelen(weken, bulk):
    """Zet bulkproducten alleen op het lijstje in de week dat ze opraken.

    Een zak van een kilo gaat meerdere weken mee, dus die hoort niet elke week
    op de lijst. Per product loopt de voorraad door over de negen weken: pas
    als er te weinig in huis is komen er zoveel verpakkingen bij als nodig.
    """
    voorraad = {}
    for w in weken:
        houden = []
        for i in w["items"]:
            regel = bulk.get(i["n"])
            nodig = naar_basis(i["q"], i["eh"]) if regel else None
            if not regel or nodig is None:
                houden.append(i)
                continue

            verpakking, eenheid, enkel, meervoud = regel
            hebben = voorraad.get(i["n"], 0.0)
            aantal = 0
            if nodig > hebben + 1e-9:
                aantal = int(-(-(nodig - hebben) // verpakking))  # naar boven af
                hebben += aantal * verpakking
            voorraad[i["n"]] = hebben - nodig

            if aantal:
                i["q"] = aantal
                i["eh"] = enkel if aantal == 1 else meervoud
                houden.append(i)
        w["items"] = houden


def strip_code(naam):
    """'D14 - Shoarma met pita's' -> 'shoarma met pita's'; zelfde regel als de app."""
    return re.sub(r"^[^-]{1,16}\s-\s", "", str(naam or "")).strip().lower()


def gebruik_per_dag(dagen, recept, hernoem):
    """Per week en ingredient: hoeveel gram er op elke weekdag gebruikt wordt.

    Nodig om te bepalen wat je zondag al in huis moet hebben en wat pas na
    woensdag nodig is. De aanvulblokken hangen niet aan een recept maar zijn
    vrije tekst ('banaan 295g + pindakaas (Calve) 10g'), die parsen we.
    """
    idx = {strip_code(r["naam"]): k for k, r in recept.items()}
    gebruik = {}
    for dag in dagen:
        week = (dag["d"] - 1) // 7
        weekdag = (dag["d"] - 1) % 7  # 0 = maandag
        for blok in dag["blokken"]:
            rid = idx.get(strip_code(blok["wat"]))
            if rid:
                paren = [(i["n"], i.get("g", 0)) for i in recept[rid]["ing"]]
            else:
                paren = []
                for deel in str(blok["wat"]).split(" + "):
                    m = re.match(r"^(.*?)\s+([\d.,]+)\s*g$", deel.strip())
                    if m:
                        paren.append((hernoem.get(m.group(1), m.group(1)),
                                      float(m.group(2).replace(",", "."))))
            for naam, gram in paren:
                naam = hernoem.get(naam, naam)
                rij = gebruik.setdefault(week, {}).setdefault(naam, [0.0] * 7)
                rij[weekdag] += gram
    return gebruik


def deel_laat(gebruik, week, naam):
    """Welk deel van een ingredient pas vanaf donderdag nodig is."""
    rij = (gebruik.get(week) or {}).get(naam)
    if not rij:
        return None  # niet te bepalen; dan gaat het gewoon mee op zondag
    totaal = sum(rij)
    return (sum(rij[3:]) / totaal) if totaal else 0.0


def maat_tekst(hoeveelheid, eenheid):
    """500 g leest prettiger dan 0,5 kg; en Nederlands schrijft een komma."""
    if eenheid == "kg" and hoeveelheid < 1:
        return "%g g" % round(hoeveelheid * 1000)
    return ("%s %s" % (("%g" % round(hoeveelheid, 2)).replace(".", ","), eenheid))


def voorraadkast(weken, bulk, prijzen):
    """Wat je van elk bulkproduct nodig hebt over de hele 60 dagen.

    Draait VOOR bulk_verdelen, zolang de weekregels nog in gram en kilo staan.
    """
    nodig = {}
    for w in weken:
        for i in w["items"]:
            if i["n"] in bulk:
                hoev = naar_basis(i["q"], i["eh"])
                if hoev is not None:
                    nodig[i["n"]] = nodig.get(i["n"], 0.0) + hoev

    rijen = []
    for naam, totaal in sorted(nodig.items()):
        verpakking, eenheid, enkel, meervoud = bulk[naam]
        prijs, winkel, _bio, _opm = prijzen.get(naam, [0, "Albert Heijn", 0, ""])
        aantal = int(-(-totaal // verpakking))  # naar boven af
        rijen.append({
            "n": naam,
            "winkel": winkel,
            "maat": maat_tekst(verpakking, eenheid),
            "aantal": aantal,
            "eh": enkel if aantal == 1 else meervoud,
            "nodig": maat_tekst(round(totaal, 2), eenheid),
            "stuk": round(prijs * verpakking, 2),
            "totaal": round(prijs * verpakking * aantal, 2),
        })
    rijen.sort(key=lambda r: -r["totaal"])
    return rijen


def bulk_prijzen(prijzen, bulk):
    """Bij bulk rekent de app per verpakking, dus de prijs mee omrekenen."""
    for naam, (verpakking, eenheid, enkel, _) in bulk.items():
        if naam not in prijzen:
            continue
        prijs, groep, bio, winkel = prijzen[naam]
        maat = "%s per %s" % (maat_tekst(verpakking, eenheid), enkel)
        # zonder winkelnotitie geen losse scheidingspunt
        prijzen[naam] = [round(prijs * verpakking, 2), groep, bio,
                         "%s · %s" % (winkel, maat) if winkel else maat]


def rondes_indelen(weken, gebruik, kort, tweewekelijks):
    """Verdeelt elke week over de zondagronde en de woensdagronde.

    Zondag haal je alles wat de week uitzingt, plus alle verse groente. Wat
    kort houdbaar is en pas na woensdag op tafel komt, koop je woensdag. Vlees
    van de lokale boer gaat de vriezer in en kan per twee weken.
    """
    for nr, w in enumerate(weken):
        nieuw = []
        oneven = w["week"] % 2 == 1
        volgende = weken[nr + 1] if nr + 1 < len(weken) else None
        # wat de volgende week aan vlees nodig heeft, koop je nu mee
        mee = {}
        if oneven and volgende:
            for x in volgende["items"]:
                if x["n"] in tweewekelijks:
                    mee[(x["n"], x["eh"])] = x

        for i in w["items"]:
            i["ronde"] = "zo"

            if i["n"] in tweewekelijks:
                if not oneven:
                    continue  # is de week ervoor al ingeslagen
                extra = mee.pop((i["n"], i["eh"]), None)
                if extra:
                    i["q"] = round(i["q"] + extra["q"], 1)
                    i["opm"] = "voor 2 weken, vriezer"
                nieuw.append(i)
                continue

            fractie = deel_laat(gebruik, nr, i["n"]) if i["n"] in kort else None
            if fractie is None or fractie <= 0.01:
                nieuw.append(i)
                continue
            if fractie >= 0.99:
                i["ronde"] = "wo"
                nieuw.append(i)
                continue

            # deels vroeg, deels laat: splitsen over de twee rondes
            laat = dict(i, q=round(i["q"] * fractie, 1), ronde="wo")
            vroeg = dict(i, q=round(i["q"] * (1 - fractie), 1), ronde="zo")
            nieuw.extend([vroeg, laat])

        # vlees dat alleen de volgende week voorkomt, hoort er nu ook bij
        for x in mee.values():
            nieuw.append(dict(x, ronde="zo", opm="voor volgende week, vriezer"))

        w["items"] = nieuw


def komma_getallen(weken):
    """Hoeveelheden als Nederlandse tekst: 1,6 liter in plaats van 1.6 liter.

    Mag als laatste stap: de app leest ze met parseFloat na een replace van de
    komma, dus rekenen blijft werken.
    """
    for w in weken:
        for i in w["items"]:
            getal = float(i["q"])
            i["q"] = (str(int(getal)) if getal == int(getal)
                      else ("%g" % getal).replace(".", ","))


def js(naam, data):
    return "export const %s = %s;\n" % (naam, json.dumps(data, ensure_ascii=False, indent=1))


def main():
    if len(sys.argv) < 2:
        sys.exit("gebruik: build-data.py <xlsx>")
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)

    recept = {}
    for blad in ("Diners", "Lunches", "Vaste modules"):
        recept.update(lees_recepten(wb[blad]))

    dagen = lees_dagen(wb)
    weken = lees_boodschappen(wb["Boodschappenlijst"])
    supplementen, aandacht = lees_supplementen(wb["Supplementen"])
    micros = lees_micros(wb["Micronutrienten"])

    winkels = json.loads(WINKELS.read_text(encoding="utf-8"))
    hernoem = winkels["hernoem"]
    hernoem_items(weken, hernoem)
    hernoem_recepten(recept, hernoem)

    # bereidingswijze en, als er een foto klaarstaat, het pad ernaartoe
    bereiding = json.loads(BEREIDING.read_text(encoding="utf-8"))
    for k, r in recept.items():
        r["stappen"] = bereiding.get(k, [])
        foto = "fotos/%s.jpg" % k
        r["foto"] = foto if (ROOT / foto).exists() else ""
    zonder = sorted(k for k, r in recept.items() if not r["stappen"])
    onbekend_recept = sorted(set(bereiding) - set(recept))

    gebruik = gebruik_per_dag(dagen, recept, hernoem)
    kast = voorraadkast(weken, winkels["bulk"], winkels["prijzen"])
    bulk_verdelen(weken, winkels["bulk"])
    bulk_prijzen(winkels["prijzen"], winkels["bulk"])

    cats = {hernoem.get(k, k): v for k, v in winkels["categorieen"].items()}
    for w in weken:
        for i in w["items"]:
            i["cat"] = cats.get(i["n"], "Overig")

    kort = set(winkels["kort_houdbaar"])
    tweewekelijks = {n for n, p in winkels["prijzen"].items()
                     if p[1] == winkels["tweewekelijks_winkel"]}
    rondes_indelen(weken, gebruik, kort, tweewekelijks)

    komma_getallen(weken)

    (ROOT / "schemaData.js").write_text(
        "// GEGENEREERD door tools/build-data.py — niet met de hand aanpassen.\n"
        + js("dagen", dagen) + js("weken", weken) + js("recept", recept)
        + js("supplementen", supplementen) + js("aandacht", aandacht)
        + js("voorraadkast", kast)
        + js("micros", micros),
        encoding="utf-8")

    (ROOT / "winkelData.js").write_text(
        "// GEGENEREERD door tools/build-data.py uit tools/winkels.json.\n"
        + js("prijzen", winkels["prijzen"]) + js("winkels", winkels["groepen"]),
        encoding="utf-8")

    onbekend = sorted({i["n"] for w in weken for i in w["items"]
                       if i["n"] not in winkels["prijzen"]})
    print("dagen: %d | recepten: %d | weken: %d | supplementen: %d"
          % (len(dagen), len(recept), len(weken), len(supplementen)))
    if onbekend:
        print("LET OP, geen prijs/winkel voor: " + ", ".join(onbekend))
    if zonder:
        print("LET OP, geen bereiding voor: " + ", ".join(zonder))
    if onbekend_recept:
        print("LET OP, bereiding voor onbekend recept: " + ", ".join(onbekend_recept))
    print("voorraadkast: %d producten, samen EUR %.2f"
          % (len(kast), sum(r["totaal"] for r in kast)))
    print("met foto: %d van de %d"
          % (sum(1 for r in recept.values() if r["foto"]), len(recept)))


if __name__ == "__main__":
    main()
